"""
Project Title: AI Job Market Analysis Pipeline
Target Role: Analytics and Insights intern (RBC Portfolio Project)

Description:
This script executes a full end-to-end data analysis pipeline on the `ai_job_market_insights.csv` dataset.
It is designed to demonstrate proficiency in Data Cleaning, SQL via SQLite, Python visualization (Plotly),
and Excel formatting (Pandas/openpyxl). This project reverse-engineers a Tableau dashboard layout conceptually 
and adds unique analytical enhancements.

Key Features & Findings:
- Profiling and cleaning an AI job market dataset (handling nulls, formatting).
- Constructing an SQLite database and executing 8 analytical queries, exporting results.
- Generating automated interactive/static charts mimicking Tableau styling using Plotly.
- Summarizing data into a heavily formatted, multi-sheet Excel workbook.
"""

import pandas as pd
import sqlite3
import os
import plotly.express as px
from openpyxl.styles import Font, PatternFill
from tabulate import tabulate

# Determine base directory
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def create_directories():
    """Creates necessary output directories if they don't exist."""
    os.makedirs(os.path.join(BASE_DIR, 'sql_outputs'), exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, 'charts'), exist_ok=True)

def step1_data_cleaning_and_profiling():
    """
    Step 1: Load, Profile, Clean, and Feature Engineer the dataset.
    """
    print("="*50)
    print("STEP 1: Data Cleaning & Profiling")
    print("="*50)
    
    # Load dataset
    csv_path = os.path.join(BASE_DIR, 'ai_job_market_insights.csv')
    df = pd.read_csv(csv_path)
    
    # 1. Profiling
    print(f"Shape: {df.shape}")
    print("\nData Types:\n", df.dtypes)
    print("\nNull Counts:\n", df.isnull().sum())
    
    print("\nSalary USD Summary Stats:\n", df['Salary_USD'].describe())
    
    print("\nCategorical Value Counts:")
    cat_cols = df.select_dtypes(include=['object']).columns
    for col in cat_cols:
        print(f"\n--- {col} ---")
        print(df[col].value_counts().head(5)) # Head 5 to keep logs clean
        
    # 2. Cleaning
    df = df.drop_duplicates()
    
    # Standardize string casing
    for col in cat_cols:
        df[col] = df[col].astype(str).str.title().str.strip()
    
    # 3. Handle any potential nulls (none found in profile, but good practice)
    df = df.fillna('Unknown')
    
    
    # Export Cleaned Data
    out_path = os.path.join(BASE_DIR, 'ai_jobs_cleaned.csv')
    df.to_csv(out_path, index=False)
    print("\nSuccessfully exported 'ai_jobs_cleaned.csv'.")
    
    return df

def step2_sql_analysis(df):
    """
    Step 2: Load cleaned data into SQLite and run 8 analytical queries.
    """
    print("\n" + "="*50)
    print("STEP 2: SQL Analysis via SQLite")
    print("="*50)
    
    # Create SQLite database and table
    db_path = os.path.join(BASE_DIR, 'ai_jobs.db')
    conn = sqlite3.connect(db_path)
    df.to_sql('ai_jobs', conn, if_exists='replace', index=False)
    
    queries = {
        "1_avg_salary_by_industry": """
            SELECT Industry, ROUND(AVG(Salary_USD), 2) as Avg_Salary
            FROM ai_jobs
            GROUP BY Industry
            ORDER BY Avg_Salary DESC;
        """,
        "2_job_count_and_salary_by_risk": """
            SELECT Automation_Risk, COUNT(*) as Job_Count, ROUND(AVG(Salary_USD), 2) as Avg_Salary
            FROM ai_jobs
            GROUP BY Automation_Risk
            ORDER BY Job_Count DESC;
        """,
        "3_top_10_required_skills": """
            SELECT Required_Skills, COUNT(*) as Skill_Frequency
            FROM ai_jobs
            GROUP BY Required_Skills
            ORDER BY Skill_Frequency DESC
            LIMIT 10;
        """,
        "4_avg_salary_by_adoption": """
            SELECT AI_Adoption_Level, ROUND(AVG(Salary_USD), 2) as Avg_Salary
            FROM ai_jobs
            GROUP BY AI_Adoption_Level
            ORDER BY Avg_Salary DESC;
        """,
        "5_company_size_distribution": """
            SELECT Company_Size, COUNT(*) as Company_Count
            FROM ai_jobs
            GROUP BY Company_Size
            ORDER BY Company_Count DESC;
        """,
        "6_avg_salary_by_location_top10": """
            SELECT Location, ROUND(AVG(Salary_USD), 2) as Avg_Salary
            FROM ai_jobs
            GROUP BY Location
            ORDER BY Avg_Salary DESC
            LIMIT 10;
        """,
        "7_automation_risk_by_industry": """
            SELECT Industry, Automation_Risk, COUNT(*) as Risk_Count
            FROM ai_jobs
            GROUP BY Industry, Automation_Risk
            ORDER BY Industry, Risk_Count DESC;
        """,
        "8_top_10_job_titles": """
            SELECT Job_Title, COUNT(*) as Job_Count
            FROM ai_jobs
            GROUP BY Job_Title
            ORDER BY Job_Count DESC
            LIMIT 10;
        """
    }
    
    for name, query in queries.items():
        # Execute query into pandas DataFrame
        res_df = pd.read_sql(query, conn)
        
        # Save to csv
        out_csv = os.path.join(BASE_DIR, 'sql_outputs', f'{name}.csv')
        res_df.to_csv(out_csv, index=False)
        
        # Print tabulate format
        print(f"\n--- Query: {name} ---")
        print(tabulate(res_df, headers='keys', tablefmt='psql', showindex=False))
        
    conn.close()

def step3_visualizations(df):
    """
    Step 3: Create, format, and save Plotly visualizations reproducing and enhancing the dashboard.
    """
    print("\n" + "="*50)
    print("STEP 3: Python Visualizations (Plotly)")
    print("="*50)
    
    # Reusable style settings mimicking Tableau Beige/Professional themes
    bg_color = '#F7F6F2'
    layout_args = dict(paper_bgcolor=bg_color, plot_bgcolor=bg_color, font=dict(family="Arial", size=12))

    # 1. Avg salary by Industry (horizontal bar)
    ind_sal = df.groupby('Industry')['Salary_USD'].mean().reset_index().sort_values('Salary_USD', ascending=True)
    fig1 = px.bar(ind_sal, x='Salary_USD', y='Industry', orientation='h', 
                  title="1. Average Salary by Industry", color_discrete_sequence=['#4E79A7'])
    fig1.update_layout(**layout_args)
    fig1.write_image(os.path.join(BASE_DIR, 'charts', '1_avg_salary_industry.png'), width=800, height=500)
    
    # 2. Automation Risk distribution (donut chart)
    risk_dist = df['Automation_Risk'].value_counts().reset_index()
    risk_dist.columns = ['Automation_Risk', 'Count']
    fig2 = px.pie(risk_dist, names='Automation_Risk', values='Count', hole=0.4, 
                  title="2. Automation Risk Distribution",
                  color_discrete_sequence=['#F28E2B', '#E15759', '#76B7B2'])
    fig2.update_layout(**layout_args)
    fig2.write_image(os.path.join(BASE_DIR, 'charts', '2_automation_risk_donut.png'), width=600, height=500)
    
    # 3. AI Adoption Level vs Avg Salary (grouped bar)
    adopt_sal = df.groupby('AI_Adoption_Level')['Salary_USD'].mean().reset_index().sort_values('Salary_USD', ascending=False)
    fig3 = px.bar(adopt_sal, x='AI_Adoption_Level', y='Salary_USD', color='AI_Adoption_Level',
                  title="3. AI Adoption Level vs Avg Salary", text_auto='.2s')
    fig3.update_layout(**layout_args)
    fig3.write_image(os.path.join(BASE_DIR, 'charts', '3_adoption_vs_salary.png'), width=600, height=500)
    
    # 4. Top 10 Skills frequency (bar chart)
    skills = df['Required_Skills'].value_counts().head(10).reset_index()
    skills.columns = ['Skill', 'Count']
    fig4 = px.bar(skills, x='Skill', y='Count', title="4. Top 10 Required Skills Frequency", 
                  color='Count', color_continuous_scale='Greens')
    fig4.update_layout(**layout_args)
    fig4.write_image(os.path.join(BASE_DIR, 'charts', '4_top_10_skills.png'), width=800, height=500)
    
    # 5. Salary distribution by Automation Risk (box plot) - ENHANCEMENT
    fig5 = px.box(df, x='Automation_Risk', y='Salary_USD', color='Automation_Risk', 
                  title="5. Salary Distribution by Automation Risk (Enhancement)")
    fig5.update_layout(**layout_args)
    fig5.write_image(os.path.join(BASE_DIR, 'charts', '5_salary_distribution_risk.png'), width=700, height=500)

    # 6. Company Size vs Automation Risk (100% stacked bar) - ENHANCEMENT
    comp_risk = df.groupby(['Company_Size', 'Automation_Risk']).size().reset_index(name='Count')
    comp_risk['Percentage'] = comp_risk.groupby('Company_Size')['Count'].transform(lambda x: x / x.sum() * 100)
    fig6 = px.bar(comp_risk, x='Company_Size', y='Percentage', color='Automation_Risk', 
                  barmode='stack', title="6. Company Size vs Automation Risk (100% Stacked)")
    fig6.update_layout(**layout_args, yaxis_title="Percentage (%)")
    fig6.write_image(os.path.join(BASE_DIR, 'charts', '6_company_size_risk_stacked.png'), width=700, height=500)
    
    # 7. Avg Salary by Location top 10 (bar chart)
    loc_sal = df.groupby('Location')['Salary_USD'].mean().reset_index().sort_values('Salary_USD', ascending=False).head(10)
    fig7 = px.bar(loc_sal, x='Location', y='Salary_USD', title="7. Avg Salary by Location (Top 10)",
                  color_discrete_sequence=['#59A14F'])
    fig7.update_layout(**layout_args)
    fig7.write_image(os.path.join(BASE_DIR, 'charts', '7_avg_salary_location.png'), width=800, height=500)
    
    # 8. KPI Summary
    total_jobs = len(df)
    avg_salary_all = df['Salary_USD'].mean()
    most_common_skill = df['Required_Skills'].mode()[0]
    highest_risk_industry = df[df['Automation_Risk'] == 'High']['Industry'].value_counts().idxmax()
    
    print("\n--- KPI Summary ---")
    print(f"Total Jobs: {total_jobs}")
    print(f"Average Salary: ${avg_salary_all:,.2f}")
    print(f"Most Common Skill: {most_common_skill}")
    print(f"Industry with Highest 'High Risk' Automation: {highest_risk_industry}")
    
    print("\nSaved 7 charts to '/charts/' directory.")

def step4_excel_summary(df):
    """
    Step 4: Create a formatted Excel workbook with Cleaned Data, Pivot tables, and Top 10 lists.
    """
    print("\n" + "="*50)
    print("STEP 4: Excel Summary Generation")
    print("="*50)
    
    excel_file = os.path.join(BASE_DIR, "ai_jobs_summary.xlsx")
    
    # Create high risk indicator
    df['Is_High_Risk'] = (df['Automation_Risk'] == 'High').astype(int)
    
    pivot_df = df.groupby('Industry').agg(
        Job_Count=('Job_Title', 'count'),
        Avg_Salary=('Salary_USD', 'mean'),
        High_Risk_Job_Count=('Is_High_Risk', 'sum')
    ).reset_index()
    
    pivot_df['High_Risk_Pct'] = (pivot_df['High_Risk_Job_Count'] / pivot_df['Job_Count']) * 100
    pivot_df['Avg_Salary'] = pivot_df['Avg_Salary'].round(2)
    pivot_df['High_Risk_Pct'] = pivot_df['High_Risk_Pct'].round(1)
    # Drop intermediate total
    pivot_df = pivot_df.drop(columns=['High_Risk_Job_Count'])
    
    # Sheet 3 Prep: Top 10 skills
    skills_df = df['Required_Skills'].value_counts().head(10).reset_index()
    skills_df.columns = ['Required_Skill', 'Job_Count']
    
    # Write to Excel
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        clean_out = df.drop(columns=['Is_High_Risk'], errors='ignore')
        clean_out.to_excel(writer, sheet_name='Cleaned_Data', index=False)
        pivot_df.to_excel(writer, sheet_name='Industry_Summary', index=False)
        skills_df.to_excel(writer, sheet_name='Top_10_Skills', index=False)
        
        # Apply formatting
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True)
        
        for sheetname in writer.sheets:
            ws = writer.sheets[sheetname]
            
            # Format Headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
                
            # Number formatting for Salary
            if sheetname == 'Cleaned_Data':
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value > 10000:
                            cell.number_format = '"$"#,##0.00'
                            
            if sheetname == 'Industry_Summary':
                for row in ws.iter_rows(min_row=2):
                    row[2].number_format = '"$"#,##0.00'
                    row[3].number_format = '0.0"%"'
                    
    print(f"Generated successfully: {excel_file}")

def step5_documentation():
    """
    Step 5: Final output summary.
    """
    print("\n" + "="*50)
    print("STEP 5: Project Summary & Outputs")
    print("="*50)
    print("All tasks completed successfully. Following files generated:\n")
    print(f"  1. ai_jobs_cleaned.csv (Cleaned Dataset)")
    print(f"  2. ai_jobs.db (SQLite Database)")
    print(f"  3. sql_outputs/ (Contains 8 CSV query results)")
    print(f"  4. charts/ (Contains 7 Plotly PNG visualizations)")
    print(f"  5. ai_jobs_summary.xlsx (Formatted multi-sheet Excel Workbook)")

if __name__ == "__main__":
    create_directories()
    df_clean = step1_data_cleaning_and_profiling()
    step2_sql_analysis(df_clean)
    step3_visualizations(df_clean)
    step4_excel_summary(df_clean)
    step5_documentation()
    print("\nPipeline Execution Complete!")
